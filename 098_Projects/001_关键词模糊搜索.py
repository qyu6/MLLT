# Databricks notebook source
# MAGIC %md
# MAGIC 本地文件关键词模糊匹配搜索并高亮

# COMMAND ----------

df = spark.read.csv('/FileStore/tables/SourceDB.csv', header="true", inferSchema="true")

print(df.dtypes)
df.select('Keywords','Content').show()

# COMMAND ----------

df1 = df.toPandas()
df1

# COMMAND ----------

# 导入相关open source library
from os.path import exists 
import openpyxl
import os
import pandas as pd 
import re
from collections import Counter

# COMMAND ----------

# 读取本地excel数据源(作为sourceDB)，设置一些前端搜索和展示信息
# 注：该demo在databricks中不能直接运行。1.从DBFS在中读取excel的部分没有调整；2.databricks前端不支持input func，不能输入关键词。(将.ipynb文件和SourceDB .xlsx文件放在本地相同路径下，直接打开即可运行。demo选择excel，而不是csv作为数据源测试，是因为csv的格式有时候会错位，不稳定。

pd.set_option('display.max_colwidth', None)
result ='searchResult.csv' 
if exists(result):
    os.remove(result)

# {create file. add header}
wbResult=openpyxl.Workbook() 
wsResult=wbResult.worksheets[0] 
wsResult.append(['result'])

# {read file.}
wb=openpyxl.load_workbook('./SourceDB.xlsx') 
input_excel='./SourceDB.xlsx' 
data=pd.read_excel(input_excel) 


ws=wb.worksheets[0]

# fill empty cell with *
for k in range(1,ws.max_column+1):	# column scan	
    for i in range(1,ws.max_row+1): # row scan
        if ws.cell(row=i, column=k).value is None:
            ws.cell(i,k,'****')	
            
            
            
input_word = input("<输入模糊搜素关键词>:").strip().lower() 
print(type(input_word))
input_word_exist = re.sub(u"([^\u4e00-\u9fa5\u0030-\u0039\u0041-\u005a\u0061-\u007a])","",input_word)
input_word = input_word.split()


#combine all values into result_list 
result_list = []
for index, row in enumerate(ws.rows):#skip header
    if index == 0:
        continue
    rs_list=list(map(lambda cell:cell.value,row))
    list_str ="".join('%s' %id for id in rs_list).replace("\n"," ").replace("\t",' ').replace("\r"," "). lower() 
    result_list.append([index,list_str])
#    result list.append(list(map(lambda item: list("".join(item)), result list)))

def search_onebyone(inpust_word_exist, input_word_list, result_list):
    new_list = [] 
    dict_list =[]
    new_list_count = []

    # accuracte matching
    for i in range(len(result_list)):
        for m in input_word_list:
            pattern=m
            regex=re.compile(pattern)
            nz=regex.search(result_list[i][1]) 
            if nz:
                new_list.append([len(nz.group()),nz.start(),result_list[i][0]-1]) 
                new_list_count.append(result_list[i][0]-1)
    # sort searching results
    new_list=sorted(new_list)
    new_index =[x for _,_,x in new_list]
    new_index=sorted(set(new_index),key=new_index.index) 
    # print(new list)


    # Count occurrence. Output when all input words are shown
    dict_list.append([k for k,v in Counter(new_list_count).items() if v== len(input_word_list)]) 
    for m in dict_list:
        result_index = m

    temp = [j for j in new_index if j in result_index] 
    return temp

result = search_onebyone(input_word_exist, input_word,result_list)
# print(result)
# print(data.loc[(x for x in result)])
data.loc[(x for x in result)].to_csv('SearchResult.csv', encoding = 'utf_8_sig')

# **************************************************************************************# highlight matching keywords in searching results. 
dftest = data.loc[(x for x in result)] 
dftest
input_word

import re
from IPython.display import HTML

def display_highlighted_words(df, keywords):
    head = """
    <table>
        <thead>
            """+\
            "".join(["<th> %s </th>" % c for c in df.columns])\
            +"""
        </thead>
    <tbody>"""

    for i,r in df.iterrows():
        row = "<tr>"
        for c in df.columns:
            matches = []
            for k in keywords:
                for match in re.finditer(k, str(r[c])):
                    matches.append(match)



            # reverse sorting
            matches = sorted(matches, key=lambda x:x.start(),reverse=True)

            # building HTML row 
            cell = str(r[c])

            #	print(cell)	
            for match in matches:
                cell=cell[:match.start()] +\
                    "<span style='background-color:yellow;color:red'> %s </span>" % cell[match.start():match.end()]+\
                    cell[match.end():] 

            row += "<td> %s </td>" % cell 
            # print(row) 
            
        row += "</tr>"
            # print(row) 
        
        head += row
        # 	print(head)	
    head += "</tbody></table>"
    
    display(HTML(head))


#原始英文单次 input word #list
#英文全部大写
temp=''.join(input_word) 
output1=temp.upper()
output1a=output1.split()

# 英文首字母大写
output2 = []
for i in input_word:
    temp = i.capitalize()
    output2.append(temp)

# 英文首字母消息，其他大写
temp = ''.join(input_word)
output3 = []
for i in input_word:
    temp1 = i[0] + i[1:].upper()
    output3.append(temp1)

final_key = set(input_word + output1a + output2 + output3)
print(final_key)


# 启动
display_highlighted_words(dftest, final_key)